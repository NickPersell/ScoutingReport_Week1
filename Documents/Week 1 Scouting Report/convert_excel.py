import pandas as pd
import json
import openpyxl
from openpyxl.styles import PatternFill, Font

def excel_to_html_with_formatting(sheet_name, excel_file):
    # Load the workbook and sheet to access formatting
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheet_name]
    
    # Read the data into a pandas DataFrame
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    
    # Convert to HTML with basic styling
    html = df.to_html(
        index=False,
        classes='table table-bordered',
        escape=False,
        border=0
    )
    
    # Parse the HTML to add custom styles based on Excel formatting
    rows = html.split('<tr>')
    styled_rows = []
    
    for row_idx, row in enumerate(rows):
        cells = row.split('<td>')
        styled_cells = [cells[0]]  # Keep the <tr> tag
        
        for col_idx, cell in enumerate(cells[1:], start=1):
            # Get the corresponding Excel cell (row_idx + 2 because of header and 0-based indexing)
            excel_cell = sheet.cell(row=row_idx + 2, column=col_idx)
            
            # Check for background color
            bg_color = ''
            if excel_cell.fill and excel_cell.fill.fgColor and excel_cell.fill.fgColor.type == 'rgb':
                bg_color = excel_cell.fill.fgColor.rgb[2:]  # Skip 'FF' prefix
                bg_color = f'background-color: #{bg_color};'
            
            # Check for bold text
            font_weight = 'font-weight: bold;' if excel_cell.font and excel_cell.font.bold else ''
            
            # Apply styles to the cell
            styled_cell = f'<td style="padding: 8px; {bg_color} {font_weight}">{cell.split("</td>")[0]}</td>'
            styled_cells.append(styled_cell)
        
        styled_rows.append('<tr>' + ''.join(styled_cells))
    
    # Reconstruct the HTML
    html = ''.join(styled_rows)
    
    # Style the headers
    html = html.replace('<th>', '<th style="background-color: #f2f2f2; padding: 8px; font-weight: bold; text-align: center;">')
    
    return html

def main():
    excel_file = 'scouting_report.xlsx'
    
    # Convert each section
    data = {
        'offense': excel_to_html_with_formatting('Offense', excel_file),
        'defense': excel_to_html_with_formatting('Defense', excel_file),
        'specialTeams': excel_to_html_with_formatting('Special Teams', excel_file)
    }
    
    # Save to JSON
    with open('data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

if __name__ == "__main__":
    main()